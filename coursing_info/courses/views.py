from django.http import HttpResponseRedirect
from django.contrib.auth import login,authenticate,logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render
from .models import course
from openpyxl.descriptors.base import Length
from openpyxl import Workbook
from openpyxl import load_workbook,worksheet
from django.contrib.auth.models import User
import os


# Create your views here.
@login_required(login_url='login')
def index(request):
    return redirect('home')

def courses_show(request,*args,**kwargs):
    course_id=kwargs['course_id']
    course_learning_manager=kwargs['course_learning_manager']
    selected_course=course.objects.filter(course_id=course_id,course_learning_manager=course_learning_manager).first()
    selected_course.courses_title.encode(encoding = 'UTF-8', errors = 'strict')
    course_title = list(selected_course.courses_title.split(","))
    course_mashaghel_link=list(selected_course.course_mashaghel_link.split(","))
    course_time=list(selected_course.course_time.split(","))
    if(selected_course.courses_teacher_name!=None):
        courses_teacher_name=list(selected_course.courses_teacher_name.split(","))
    course_teacher_rezume=list(selected_course.course_teacher_rezume.split(","))
    course_id_list=list(course_id.split(","))

    i=0
    specific_detailes=[]
    specific_detail={}
    print(len(course_title))
    while i<=len(course_id_list)-1:
        print(i)
        specific_detail={
            'course_title':course_title[i],
            'course_mashaghel_link':course_mashaghel_link[i],
            'course_time':course_time[i],
            'course_id_list':course_id_list[i],
            'course_teacher_rezume':course_teacher_rezume[i],
        }
        specific_detailes.append(specific_detail)       
        i=i+1

    context={
        'course':selected_course,
        'specific_detailes':specific_detailes
    }
    return render(request,'course.html',context)


@login_required(login_url='login')
def upload_course(request):
    context={}
    if (request.POST or None):
        file_name=request.POST['file_name']     
        if os.path.exists(f'media/{file_name}.xlsx')==True:
            wb=Workbook()
            wb = load_workbook(filename = f'media/{file_name}.xlsx')
            sheet = wb['Sheet1']
            sheetlenght=len(sheet['A'])
            l=2
            course_sections=[]
            course_mashaghel_link=[]
            course_time=[]
            course_targets=[]
            courses_teacher_name=[]
            course_id=[]
            course_teacher_rezume=[]
            courses_title=[]
            while l<=sheetlenght:
                courses_online=sheet['C'+str(l)].value
                course_video=sheet['D'+str(l)].value
                courses_date=sheet['F'+str(l)].value
                course_learning_manager=sheet['M'+str(l)].value
                course_contact_number_1=sheet['N'+str(l)].value
                course_contact_number_2=sheet['O'+str(l)].value
                course_whatsapp_number=sheet['P'+str(l)].value
                course_category=sheet['Q'+str(l)].value
                course_image=sheet['R'+str(l)].value

                if (sheet['A'+ str(l)].value==sheet['A'+str(l+1)].value):
                    course_mashaghel_link.append(sheet['E'+str(l)].value)
                    course_sections.append(sheet['L'+str(l)].value)
                    course_time.append(sheet['J'+str(l)].value)
                    course_targets.append(sheet['K'+str(l)].value)
                    courses_teacher_name.append(sheet['G'+str(l)].value)
                    course_id.append(sheet['I'+str(l)].value)
                    course_teacher_rezume.append(sheet['H'+str(l)].value)
                    courses_title.append(sheet['B'+str(l)].value)
                    l=l+1
                else:
                    course_mashaghel_link.append(sheet['E'+str(l)].value)
                    course_mashaghel_link=''.join(str(e+',') for e in course_mashaghel_link)
                    course_sections.append(sheet['L'+str(l)].value)
                    course_sections=''.join(str(e+',') for e in course_sections)
                    course_time.append(sheet['J'+str(l)].value)
                    course_time=str(course_time)
                    course_time=course_time.replace('[','')
                    course_time=course_time.replace(']','')
                    course_targets.append(sheet['K'+str(l)].value)
                    course_targets=''.join(str(e+',') for e in course_targets)
                    courses_teacher_name.append(sheet['G'+str(l)].value)
                    courses_teacher_name=''.join(str(e+',') for e in courses_teacher_name)
                    course_id.append(sheet['I'+str(l)].value)
                    course_id=str(course_id)
                    course_id=course_id.replace('[','')
                    course_id=course_id.replace(']','')
                    course_teacher_rezume.append(sheet['H'+str(l)].value)
                    course_teacher_rezume=''.join(str(e+',') for e in course_teacher_rezume)
                    courses_title.append(sheet['B'+str(l)].value)
                    courses_title=''.join(str(e+',') for e in courses_title)
                    # create course by l
                    createcourse=course.objects.create(
                    courses_title=courses_title,
                    courses_online=courses_online,
                    course_video=course_video,
                    course_mashaghel_link=course_mashaghel_link,
                    courses_date=courses_date,
                    courses_teacher_name=courses_teacher_name,
                    course_teacher_rezume=course_teacher_rezume,
                    course_id=course_id,
                    course_time=course_time,
                    course_targets=course_targets,
                    course_sections=course_sections,
                    course_learning_manager=course_learning_manager,
                    course_contact_number_1=course_contact_number_1,
                    course_contact_number_2=course_contact_number_2,
                    course_whatsapp_number=course_whatsapp_number,
                    course_category=course_category,
                    course_image=course_image,
                    course_writer=request.user   
            )     
                    course_sections=[]
                    course_mashaghel_link=[]
                    course_time=[]
                    course_targets=[]
                    courses_teacher_name=[]
                    course_id=[]
                    course_teacher_rezume=[]
                    courses_title=[]
                    l=l+1 
                context={
                    'msg':'فراخوان ها با موفقیت ایجاد شده اند.'
                }
                
            return redirect('showlinks')
        else:
         context={
            'msg':'فایل مورد نظر یافت نشد'
            } 

    return render(request,'uploading.html',context)


def logindef(request):
   if request.user.is_authenticated:
        return redirect('home')
   else:
    msg=''
    if(request.POST or None):
        username = request.POST['username']
        password = request.POST['password']
        User = authenticate(request, username=username, password=password)
        if User is not None:
            login(request,User)
            msg='شما با موفقیت وارد شدید'  
            return redirect('showlinks')          
        else:
            msg='نام کاربری یا رمز ورود اشتباه میباشد!'

    context={
        'msg':msg
    }
    return render(request,'login.html',context)


def exit(request):
    if request.user.is_authenticated:
        logout(request)
        print(request.user)
    return redirect('login')




def showlinks(request):
    selectallcourse=course.objects.all()
    # print(selectallcourse)
    context={
        'courses':selectallcourse
    }
    return render(request,'showlinks.html',context)